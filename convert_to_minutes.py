import openpyxl
import math
import os
from openpyxl.xml.constants import MAX_ROW
cwd = os.getcwd()
folder = os.fsencode(cwd)
filenames = []
for file in os.listdir(folder):
    filename = os.fsdecode(file)
    if filename.endswith(('.xlsx')):
        filenames.append(filename)
print(filenames)
for f in filenames:        
    wb = openpyxl.load_workbook(f)
    sheet = wb.get_active_sheet()
    toadd = 1
    if (sheet['A1'].value) == "Current Time (ms)":
        sheet.delete_rows(1,1)
        toadd = 0
    print(sheet['A1'].value)
    x = sheet['A1'].value
    print(f)
    for q in range(1, sheet.max_row + toadd):
        if type(sheet.cell(row = q, column =1).value) != int:
            break
        sheet.cell(row = q, column =1).value = (sheet.cell(row = q, column =1).value - x)
        m = (sheet.cell(row = q, column =1).value)       
        second = (m/ 1000)/60  
        y = math.modf(second)
        y = ((y[0] * 60) / 100) + y[1]
        sheet.cell(row = q, column =1).value = ((sheet.cell(row = q, column =1).value)) = y
    newname = f.replace(".xlsx", "")
    newname = newname + '_actual_time.xlsx'                                          
    wb.save(newname)
