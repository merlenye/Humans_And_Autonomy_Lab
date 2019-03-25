import openpyxl
import math
import os
from openpyxl import Workbook
from openpyxl.xml.constants import MAX_ROW
def filelist(): 
    names = []
    for file in os.listdir(folder):
        filename = os.fsdecode(file)
        if filename.endswith(('.xlsx')):
            names.append(filename)
    return names
def minutize(mili):
    mili = (mili/ 1000)/60  
    y = math.modf(mili)
    y = ((y[0] * 60) / 100) + y[1]
    return y
def makebook(filenames):
    rowcount = 0
    finalbook = Workbook()
    finalsheet = finalbook.active  
    for f in filenames:     
        wb = openpyxl.load_workbook(f)
        sheet = wb.get_active_sheet()
        print(sheet['A1'].value)
        x = sheet['A1'].value
        print(f)
        cntsup = 0
        cntins = 0
        super = 0
        insp = 0
        last = 0
        isfirst = 0
        for q in range(1, sheet.max_row + 1):
            if type(sheet.cell(row = q, column =1).value) != int:
                if last == 4:
                    insp = insp + sheet.cell(row = q-1, column = 1).value - cntins
                if last == 5:
                   super = super + sheet.cell(row = q-1, column = 1).value - cntsup
                break  
            sheet.cell(row = q, column = 1).value = sheet.cell(row = q, column = 1).value - x
            if q == (sheet.max_row):
                if last == 4:
                    insp = insp + sheet.cell(row = q, column = 1).value - cntins
                if last == 5:
                   super = super + sheet.cell(row = q, column = 1).value - cntsup 
            
                break
            if sheet.cell(row = q, column = 6).value == 4:
                if isfirst > 0:
                    super = super + sheet.cell(row = q, column = 1).value - cntsup
                    last = 4
                    print(super)
                    cntins = sheet.cell(row = q, column = 1).value
                isfirst +=1
                last = 4
            if sheet.cell(row = q, column = 6).value == 5:
                insp = insp + sheet.cell(row = q, column = 1).value - cntins
                cntsup = sheet.cell(row = q, column = 1).value
                last =5
        print(insp)
        print(super)
        rowcount+= 1
        finalsheet.cell(row = rowcount, column = 1).value = f
        finalsheet.cell(row = rowcount, column = 2).value = (minutize(insp))
        finalsheet.cell(row = rowcount, column = 3).value = (minutize(super))
    wb.save("abso.xlsx")
    finalbook.save("time_portions.xlsx")
if __name__ == '__main__':
    cwd = os.getcwd()
    folder = os.fsencode(cwd)
    filenames = filelist()
    print(filenames)
    makebook(filenames)
    
