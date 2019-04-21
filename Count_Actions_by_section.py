# Simple code to count the number of actions the pilot preformed each section of the expirement and record them
#in an excel file with 
from lib2to3.tests.pytree_idempotency import diff
'''
Created on Apr 15, 2019

@author: merlenye
'''
import openpyxl
import math
import os
from openpyxl import Workbook
from openpyxl.xml.constants import MAX_ROW
           
def filelist(): 
    names = []
    for file in os.listdir(folder):
        filename = os.fsdecode(file)
        if filename.endswith(('.xlsx')):
            names.append(filename)
    return names
def makebook(filenames):
    finalbook = Workbook()
    finalsheet = finalbook.active  
    rowcount = 1
    for f in filenames:   
        wb = openpyxl.load_workbook(f)
        sheet = wb.get_active_sheet()
        print(f)
        rowcount +=1
        col_count = 1
        counter =0
        sec_cnt = 0
        finalsheet.cell(row = rowcount, column = 1).value = f
        for q in range(1, sheet.max_row + 1): 
            print(q) 
            counter+=1
            if (sheet.cell(row = counter, column =1).value == 5555):
                col_count +=1
                finalsheet.cell(row = rowcount, column = col_count).value = sec_cnt
                sec_cnt=0
                counter+=1
            if (sheet.cell(row = counter, column =1).value == 4444):
                col_count+=1
                finalsheet.cell(row = rowcount, column = col_count).value = sec_cnt
                sec_cnt=0
                counter+=1
                break  
            if (sheet.cell(row = counter, column =6).value != 12):  
                sec_cnt+= 1
                

        finalbook.save("actions_per_section.xlsx")
if __name__ == '__main__':  
    cwd = os.getcwd()
    folder = os.fsencode(cwd)
    filenames = filelist()
    print(filenames)
    makebook(filenames)
